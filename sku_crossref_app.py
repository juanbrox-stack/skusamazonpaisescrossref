import streamlit as st
import pandas as pd
import io, re, requests
from openpyxl import load_workbook

st.set_page_config(page_title="SKU CrossRef – Amazon", layout="wide", page_icon="🔗")

# ─── SKU structure ────────────────────────────────────────────────────────────
# Jabiru ES / Turaco ES : SKU base  +  S+SKU
# FR / IT / DE          : SKU base  +  S+SKU  +  XX+SKU  (XX = FR / IT / DE)
# NL / PL / SE          : SKU base  +  S+SKU
# Prestashop            : must have every SKU that exists in any Amazon store

STORE_CONFIG = {
    # key: (label, flag, prefixes_required, file_key)
    "jabiru_es":  ("Jabiru ES",  "🇪🇸", ["", "S"],           "jabiru_es"),
    "jabiru_fr":  ("Jabiru FR",  "🇫🇷", ["", "S", "FR"],     "jabiru_fr"),
    "jabiru_it":  ("Jabiru IT",  "🇮🇹", ["", "S", "IT"],     "jabiru_it"),
    "jabiru_de":  ("Jabiru DE",  "🇩🇪", ["", "S", "DE"],     "jabiru_de"),
    "jabiru_nl":  ("Jabiru NL",  "🇳🇱", ["", "S"],           "jabiru_nl"),
    "jabiru_pl":  ("Jabiru PL",  "🇵🇱", ["", "S"],           "jabiru_pl"),
    "jabiru_se":  ("Jabiru SE",  "🇸🇪", ["", "S"],           "jabiru_se"),
    "turaco_es":  ("Turaco ES",  "🇪🇸", ["", "S"],           "turaco_es"),
    "turaco_fr":  ("Turaco FR",  "🇫🇷", ["", "S", "FR"],     "turaco_fr"),
    "turaco_it":  ("Turaco IT",  "🇮🇹", ["", "S", "IT"],     "turaco_it"),
    "turaco_de":  ("Turaco DE",  "🇩🇪", ["", "S", "DE"],     "turaco_de"),
}

AMZN_AUTO = re.compile(r'^AMZN\.', re.IGNORECASE)

# ─── Core helpers ─────────────────────────────────────────────────────────────
def extract_base(ref: str) -> str:
    ref = ref.strip().rstrip(".")
    if "_" in ref:
        m = re.search(r"([A-Z]\d{2}_)", ref.upper())
        return ref[m.start():] if m else ref
    m = re.match(r"^(DE|FR|IT|NL|PL|SE|S)([A-Z]?\d+.*)$", ref, re.IGNORECASE)
    if m:
        rest = m.group(2)
        return rest.zfill(5) if re.match(r"^\d+$", rest) else rest
    return ref.zfill(5) if re.match(r"^\d+$", ref) else ref

@st.cache_data(show_spinner=False)
def parse_listing(data: bytes, label: str) -> tuple:
    """Returns (df_active, df_all, attrs). Cached by file content hash."""
    try:
        raw = pd.read_csv(io.BytesIO(data), sep="\t", dtype=str, low_memory=False,
                          encoding="utf-8-sig", on_bad_lines="skip")
    except Exception as e:
        st.error(f"❌ {label}: {e}")
        empty = pd.DataFrame(columns=["sku","asin"])
        return empty, empty, {"label": label, "total": 0, "active": 0}

    sc = next((c for c in raw.columns if c.strip().lower() == "status"), None)
    total = len(raw)
    active_n = int((raw[sc].str.strip().str.lower() == "active").sum()) if sc else total

    def _ex(df):
        out = df[[df.columns[0], df.columns[1]]].copy()
        out.columns = ["sku","asin"]
        out["sku"] = out["sku"].astype(str).str.strip().str.upper()
        out["asin"] = out["asin"].astype(str).str.strip()
        return out[out["sku"].notna() & (out["sku"] != "") & (out["sku"] != "nan")]

    df_all = _ex(raw)
    df_active = _ex(raw[raw[sc].str.strip().str.lower() == "active"]) if sc else df_all.copy()
    return df_active, df_all, {"label": label, "total": total, "active": active_n}

@st.cache_data(show_spinner=False)
def parse_ps(data: bytes) -> pd.Series:
    df = pd.read_csv(io.BytesIO(data), sep=";", dtype=str, low_memory=False)
    col = next((c for c in df.columns if c.strip().lower() == "reference"), None)
    if not col: return pd.Series(dtype=str)
    s = df[col].dropna().str.strip()
    return s[s != ""]

def sku_set(df): return set(df["sku"].str.upper().dropna())
def sku_map(df): return df.drop_duplicates("sku").set_index("sku")["asin"].to_dict()

def jabiru_bases(jdf: pd.DataFrame) -> dict:
    """Extract unique bases from Jabiru ES active SKUs. Ignores AMZN.* auto-SKUs."""
    result = {}
    for _, row in jdf.iterrows():
        if AMZN_AUTO.match(row["sku"]): continue
        b = extract_base(row["sku"]).upper()
        if b not in result:
            result[b] = (row["sku"], row["asin"])
    return result

# ─── Cross-check functions ────────────────────────────────────────────────────
def missing_variants(j_bases: dict, target_skus: set, prefixes: list) -> pd.DataFrame:
    """
    For each Jabiru ES base, checks that ALL required prefix variants exist in target.
    Returns rows where at least one required variant is missing.
    prefixes: list of prefixes to check, e.g. ["", "S"] or ["", "S", "FR"]
    """
    rows = []
    for base, (sku_j, asin) in j_bases.items():
        missing = []
        for pfx in prefixes:
            variant = f"{pfx}{base}"
            if variant not in target_skus:
                missing.append(variant)
        if missing:
            all_variants = [f"{p}{base}" for p in prefixes]
            rows.append({
                "SKU Jabiru ES":    sku_j,
                "Base SKU":         base,
                "ASIN":             asin,
                "Faltantes":        " | ".join(missing),
                "Variantes req.":   " | ".join(all_variants),
            })
    return pd.DataFrame(rows)

def ps_missing(j_bases: dict, turaco_sets: dict, ps_refs: pd.Series) -> pd.DataFrame:
    """
    Check which SKUs (across all stores and all prefix variants) are missing in PS.
    PS must have every single SKU that any store uses.
    """
    ps_set = set(extract_base(r).upper() for r in ps_refs)
    # Also include S+base, FR+base etc. as PS references to check
    # In PS, references are stored as: 00234, S00234, FR00234, etc.
    ps_refs_upper = set(r.strip().upper() for r in ps_refs)

    rows = []
    seen = set()
    for base, (sku_j, asin) in j_bases.items():
        # All variants that should exist across all stores
        all_variants = set()
        all_variants.update([base, f"S{base}"])          # ES / NL / PL / SE
        all_variants.update([f"FR{base}", f"IT{base}", f"DE{base}"])  # international
        for v in all_variants:
            if v not in ps_refs_upper and v not in seen:
                seen.add(v)
                rows.append({
                    "Referencia faltante PS": v,
                    "Base SKU":              base,
                    "ASIN":                  asin,
                    "SKU Jabiru ES":         sku_j,
                })
    return pd.DataFrame(rows)

# ─── PS CSV generator ─────────────────────────────────────────────────────────
PS_TPL_COLS = [
    "Product ID","Active (0/1)","Name *","Categories (x,y,z...)","Price tax included",
    "Tax rules ID","Wholesale price","On sale (0/1)","Discount amount","Discount percent",
    "Discount from (yyyy-mm-dd)","Discount to (yyyy-mm-dd)","Reference #",
    "Supplier reference #","Supplier","Manufacturer","EAN13","UPC","Ecotax",
    "Width","Height","Depth","Weight","Quantity","Minimal quantity","Low stock level",
    "Visibility","Additional shipping cost","Unity","Unit price",
    "Short description","Description","Tags (x,y,z...)","Meta title","Meta keywords",
    "Meta description","URL rewritten","Text when in stock","Text when backorder allowed",
    "Available for order (0 = No, 1 = Yes)","Product available date",
    "Product creation date","Show price (0 = No, 1 = Yes)",
    "Image URLs (x,y,z...)","Image alt texts (x,y,z...)",
    "Delete existing images (0 = No, 1 = Yes)","Feature(Name:Value:Position)",
    "Available online only (0 = No, 1 = Yes)","Condition",
    "Customizable (0 = No, 1 = Yes)","Uploadable files (0 = No, 1 = Yes)",
    "Text fields (0 = No, 1 = Yes)","Out of stock","ID / Name of shop",
    "Advanced stock management","Depends On Stock","Warehouse",
]

@st.cache_data(show_spinner=False)
def build_ps_lookups(ps_sql_bytes: bytes, ps_carga_bytes: bytes | None):
    sql = pd.read_csv(io.BytesIO(ps_sql_bytes), sep=";", dtype=str, low_memory=False)
    sql["ref_key"] = sql["reference"].str.strip().str.upper()
    sql_lk = sql.drop_duplicates("ref_key").set_index("ref_key")
    carga_lk = None
    if ps_carga_bytes:
        carga = pd.read_excel(io.BytesIO(ps_carga_bytes), dtype=str)
        if "Reference #" in carga.columns:
            carga["ref_key"] = carga["Reference #"].str.strip().str.upper()
            carga_lk = carga.drop_duplicates("ref_key").set_index("ref_key")
    return sql_lk, carga_lk

def make_ps_row(ref: str, base: str, sql_lk, carga_lk) -> dict:
    sql_d   = sql_lk.loc[base]   if sql_lk   is not None and base in sql_lk.index   else None
    carga_d = carga_lk.loc[base] if carga_lk is not None and base in carga_lk.index else None
    price_ti = ""
    if carga_d is not None:
        price_ti = str(carga_d.get("Price tax included","") or "").strip()
    if not price_ti and sql_d is not None:
        try: price_ti = str(round(float(str(sql_d.get("price","0")).replace(",",".")) * 1.21, 2))
        except: pass
    r = {c: "" for c in PS_TPL_COLS}
    r.update({"Active (0/1)":"1","Reference #":ref,"Supplier reference #":ref,
               "Tax rules ID":"1","On sale (0/1)":"0","Quantity":"0",
               "Minimal quantity":"1","Visibility":"both","Text when in stock":"In Stock",
               "Available for order (0 = No, 1 = Yes)":"1","Show price (0 = No, 1 = Yes)":"1",
               "Delete existing images (0 = No, 1 = Yes)":"0",
               "Available online only (0 = No, 1 = Yes)":"1","Condition":"new",
               "Customizable (0 = No, 1 = Yes)":"0","Uploadable files (0 = No, 1 = Yes)":"0",
               "Text fields (0 = No, 1 = Yes)":"0","Out of stock":"0",
               "ID / Name of shop":"0","Advanced stock management":"0",
               "Depends On Stock":"0","Warehouse":"0"})
    if sql_d is not None:
        r["EAN13"] = str(sql_d.get("ean13","") or "").strip()
        for f in ["width","height","depth","weight"]:
            r[f.capitalize()] = str(sql_d.get(f,"1") or "1").strip() or "1"
        r["Wholesale price"] = str(sql_d.get("wholesale_price","") or "").strip()
        r["Supplier"] = r["Manufacturer"] = "Cecotec"
    if carga_d is not None:
        for field in ["Name *","Categories (x,y,z...)","Description","Short description",
                      "Image URLs (x,y,z...)","Image alt texts (x,y,z...)","Tags (x,y,z...)",
                      "EAN13","Width","Height","Depth","Weight","Wholesale price",
                      "Supplier","Manufacturer","Meta title","Meta keywords","Meta description"]:
            val = carga_d.get(field,"")
            if pd.notna(val) and str(val).strip():
                r[field] = str(val).strip()
    if price_ti: r["Price tax included"] = price_ti
    return r

def gen_ps_csv(df_missing: pd.DataFrame, sql_lk, carga_lk) -> bytes:
    rows = []
    for _, row in df_missing.iterrows():
        ref  = str(row.get("Referencia faltante PS", row.get("Faltantes",""))).strip()
        base = str(row.get("Base SKU","")).strip().upper()
        # If multiple faltantes in one row, split and create one row per variant
        for r in ref.split("|"):
            r = r.strip()
            if r:
                rows.append(make_ps_row(r, base, sql_lk, carga_lk))
    df_out = pd.DataFrame(rows, columns=PS_TPL_COLS)
    buf = io.BytesIO()
    df_out.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue()

# ─── Amazon template ──────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def parse_feed_file(data: bytes, label: str) -> dict:
    """
    Parse a Cecotec feed file (JSON or CSV/TSV) uploaded by the user.
    Extracts {mpn_upper: price_str} mapping.
    Supports:
      - JSON feed from /api/v3/doofinder/feed/  (save URL as .json in browser)
      - CSV/TSV with columns: mpn/reference/sku + price/sale_price
    """
    import json as _json
    m = {}

    def _add(mpn, price):
        mpn = str(mpn).strip().upper()
        price = str(price).strip()
        if mpn and price and price not in ("", "0", "0.0", "nan"):
            m[mpn] = price
            stripped = mpn.lstrip("0") or "0"
            if stripped != mpn:
                m[stripped] = price

    # Try JSON first
    try:
        text = data.decode("utf-8", errors="replace").strip()
        data_j = _json.loads(text)
        items = data_j if isinstance(data_j, list) else data_j.get(
            "items", data_j.get("results", data_j.get("products", [])))
        for item in items:
            mpn   = item.get("mpn") or item.get("reference") or item.get("id","")
            price = item.get("price") or item.get("sale_price","")
            _add(mpn, price)
        if m:
            return m
    except Exception:
        pass  # not JSON, try CSV

    # Try CSV/TSV
    try:
        sep = "	" if b"	" in data[:500] else ","
        df = pd.read_csv(io.BytesIO(data), sep=sep, dtype=str, low_memory=False,
                         encoding="utf-8-sig", on_bad_lines="skip")
        df.columns = [c.strip().lower() for c in df.columns]
        mpn_col   = next((c for c in df.columns if c in ("mpn","reference","sku","ref")), None)
        price_col = next((c for c in df.columns if c in ("price","sale_price","precio","pvp")), None)
        if mpn_col and price_col:
            for _, row in df.iterrows():
                _add(row[mpn_col], row[price_col])
    except Exception as e:
        return {"__error__": f"No se pudo leer el fichero de feed {label}: {e}"}

    if not m:
        return {"__error__": f"Feed {label}: no se encontraron productos (comprueba columnas mpn/reference y price)"}
    return m

def get_price(base: str, feed: dict, es_feed: dict) -> str:
    def _look(m, b):
        if not m or "__error__" in m: return ""
        for k in [b, b.lstrip("0") or "0", b.zfill(5)]:
            if k in m: return m[k]
        return ""
    return _look(feed, base) or _look(es_feed, base)

def gen_amz_template(df: pd.DataFrame, feed: dict, es_feed: dict, rate: float) -> bytes:
    tmpl = st.session_state.get("template_bytes")
    if not tmpl: raise ValueError("Sube la plantilla Amazon ES en el panel lateral.")
    wb = load_workbook(io.BytesIO(tmpl), keep_vba=True)
    ws = wb["Plantilla"]
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        for cell in row: cell.value = None
    for i, (_, row) in enumerate(df.iterrows(), start=7):
        asin = row.get("ASIN","")
        sku  = row.get("SKU Jabiru ES", row.get("Faltantes","").split("|")[0].strip())
        base = row.get("Base SKU","")
        ws.cell(i,1).value  = asin
        ws.cell(i,4).value  = "Añadir producto"
        ws.cell(i,5).value  = sku
        ws.cell(i,6).value  = asin
        ws.cell(i,12).value = "Nuevo"
        ws.cell(i,36).value = "Logística por parte del vendedor (predeterminado)"
        ws.cell(i,40).value = "Habilitado"
        p = get_price(base, feed, es_feed)
        if p:
            try: ws.cell(i,41).value = round(float(p.replace(",",".")) * rate, 2)
            except: pass
        ws.cell(i,66).value = "FBM NO HB"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── UI helpers ───────────────────────────────────────────────────────────────
def show_table(df: pd.DataFrame, dl_key: str, dl_name: str):
    if df.empty:
        st.success("✅ Sin diferencias.")
        return
    st.warning(f"⚠️ **{len(df):,}** bases con variantes faltantes.")
    q = st.text_input("🔍 Filtrar", key=f"q_{dl_key}")
    filt = df
    if q:
        mask = pd.Series(False, index=df.index)
        for c in df.columns:
            mask |= df[c].astype(str).str.contains(q, case=False, na=False)
        filt = df[mask]
    st.dataframe(filt, width="stretch", height=380)
    st.download_button("⬇️ CSV", filt.to_csv(index=False).encode("utf-8-sig"),
                       file_name=dl_name, mime="text/csv", key=f"dl_{dl_key}")

def action_buttons(df: pd.DataFrame, dl_key: str, label: str,
                   feed: dict, es_feed: dict, rate: float,
                   sql_lk, carga_lk,
                   show_amz: bool = False, show_ps: bool = True):
    """Render download action buttons below a results table."""
    if df.empty: return
    col1, col2 = st.columns(2) if (show_amz and show_ps) else (st, None)

    if show_amz:
        target = col1 if col2 else st
        with target.expander("📋 Plantilla Amazon ES"):
            if not st.session_state.get("template_bytes"):
                st.info("📤 Sube la plantilla Amazon ES en el panel lateral.")
            elif st.button(f"Generar – {label}", key=f"amz_{dl_key}"):
                with st.spinner("Generando…"):
                    try:
                        data = gen_amz_template(df, feed, es_feed, rate)
                        st.download_button("⬇️ Plantilla Amazon ES", data=data,
                                           file_name=f"amz_{dl_key}.xlsm",
                                           mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                                           key=f"amz_dl_{dl_key}")
                    except Exception as e: st.error(str(e))

    if show_ps and sql_lk is not None:
        target = col2 if col2 else st
        with target.expander("🛒 CSV carga Prestashop"):
            if not carga_lk:
                st.caption("💡 Sube Carga_PS en el lateral para enriquecer con nombre e imágenes.")
            if st.button(f"Generar CSV PS – {label}", key=f"ps_{dl_key}"):
                with st.spinner("Generando…"):
                    csv_b = gen_ps_csv(df, sql_lk, carga_lk)
                    st.download_button("⬇️ CSV carga PS", data=csv_b,
                                       file_name=f"carga_PS_{dl_key}.csv",
                                       mime="text/csv", key=f"ps_dl_{dl_key}")

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.header("📂 Archivos base (obligatorios)")
    ps_file     = st.file_uploader("📦 Prestashop BD (CSV ;)", type=["csv","txt"], key="ps")
    jabiru_file = st.file_uploader("🇪🇸 Jabiru ES",             type=["txt","tsv","csv"], key="jabiru_es")

    st.divider()
    st.header("🏪 Tiendas a comparar")
    st.caption("Activa y sube el listing de cada tienda:")

    store_files = {}
    store_enabled = {}
    OTHER_STORES = [
        ("jabiru_fr",  "Jabiru FR",  "🇫🇷"),
        ("jabiru_it",  "Jabiru IT",  "🇮🇹"),
        ("jabiru_de",  "Jabiru DE",  "🇩🇪"),
        ("jabiru_nl",  "Jabiru NL",  "🇳🇱"),
        ("jabiru_pl",  "Jabiru PL",  "🇵🇱"),
        ("jabiru_se",  "Jabiru SE",  "🇸🇪"),
        ("turaco_es",  "Turaco ES",  "🇪🇸"),
        ("turaco_fr",  "Turaco FR",  "🇫🇷"),
        ("turaco_it",  "Turaco IT",  "🇮🇹"),
        ("turaco_de",  "Turaco DE",  "🇩🇪"),
    ]
    for key, lbl, flag in OTHER_STORES:
        enabled = st.checkbox(f"{flag} {lbl}", value=False, key=f"chk_{key}")
        store_enabled[key] = enabled
        if enabled:
            f = st.file_uploader(f"↳ {lbl}", type=["txt","tsv","csv"],
                                 key=f"file_{key}", label_visibility="collapsed")
            store_files[key] = f
        else:
            store_files[key] = None

    st.divider()
    st.header("📋 Extras")
    tpl = st.file_uploader("Plantilla Amazon ES (.xlsm)", type=["xlsm","xlsx"], key="tpl")
    if tpl:
        st.session_state["template_bytes"] = tpl.read()
        st.success("✅ Plantilla lista")
    ps_carga = st.file_uploader("Carga_PS enriquecimiento (.xlsx)", type=["xlsx"], key="ps_carga")
    if ps_carga:
        st.session_state["ps_carga_bytes"] = ps_carga.read()
        st.success("✅ Carga_PS lista")

    st.divider()
    st.header("💱 Tipos de cambio")
    rate_sek = st.number_input("SEK/EUR", value=st.session_state.get("rate_sek",11.5),
                                min_value=0.01, step=0.1, key="rate_sek")
    rate_pln = st.number_input("PLN/EUR", value=st.session_state.get("rate_pln",4.25),
                                min_value=0.01, step=0.01, key="rate_pln")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
st.title("🔗 SKU CrossRef – Amazon Listings")
st.markdown("""
**Fuente de verdad:** Jabiru ES. Comprueba que todos los marketplaces y Prestashop tienen
paridad de SKUs con Jabiru ES.

| Tienda | Variantes requeridas |
|---|---|
| Jabiru ES / Turaco ES | `SKU`  `S+SKU` |
| Jabiru FR / IT / DE / Turaco FR / IT / DE | `SKU`  `S+SKU`  `XX+SKU` |
| Jabiru NL / PL / SE | `SKU`  `S+SKU` |
| Prestashop | todas las variantes de todas las tiendas |
""")

if not ps_file or not jabiru_file:
    st.info("👈 Carga el CSV de Prestashop y el listing de **Jabiru ES** para comenzar.")
    st.stop()

# ─── Load base data ───────────────────────────────────────────────────────────
ps_raw = ps_file.read()
st.session_state["ps_bytes_raw"] = ps_raw
refs = parse_ps(ps_raw)

jab_active, jab_all, jab_attrs = parse_listing(jabiru_file.read(), "Jabiru ES")
j_bases = jabiru_bases(jab_active)
jab_skus = sku_set(jab_all)

# PS lookups
sql_lk, carga_lk = build_ps_lookups(
    st.session_state["ps_bytes_raw"],
    st.session_state.get("ps_carga_bytes")
)

# Feeds — upload-based (URL fetch blocked by Streamlit Cloud network policy)
feed_maps = st.session_state.get("feed_maps", {})
with st.expander("💶 Precios desde feeds Cecotec (opcional)"):
    st.caption(
        "Descarga cada feed en tu navegador y súbelo aquí. URLs de descarga:\n"
        "- **ES**: https://cecotec.es/api/v3/doofinder/feed/?lang=es  \n"
        "- **FR**: https://storececotec.fr/api/v3/doofinder/feed/?lang=fr  \n"
        "- **IT**: https://content.storececotec.it/api/v3/doofinder/feed/?lang=it  \n"
        "- **DE**: https://storececotec.de/api/v3/doofinder/feed/?lang=de  \n"
        "Guarda cada URL como archivo (.json, .csv o .txt) y súbelo abajo."
    )
    FEED_LABELS = {"ES": "🇪🇸 Feed ES", "FR": "🇫🇷 Feed FR",
                   "IT": "🇮🇹 Feed IT", "DE": "🇩🇪 Feed DE"}
    cols = st.columns(4)
    for i, (k, lbl) in enumerate(FEED_LABELS.items()):
        with cols[i]:
            uf = st.file_uploader(lbl, type=["json","csv","txt","tsv"], key=f"feed_file_{k}")
            if uf:
                feed_data = parse_feed_file(uf.read(), k)
                if "__error__" in feed_data:
                    st.warning(feed_data["__error__"])
                else:
                    feed_maps[k] = feed_data
                    st.session_state["feed_maps"] = feed_maps
                    st.success(f"✅ {len(feed_data)//2} productos")
            elif k in feed_maps:
                st.caption(f"✅ {len(feed_maps[k])//2} cargados")

es_feed = feed_maps.get("ES", {})

st.caption(f"Jabiru ES: **{jab_attrs['active']:,}** activos / {jab_attrs['total']:,} total  "
           f"| PS referencias: **{len(refs):,}**")
st.divider()

# ─── Comparativas ─────────────────────────────────────────────────────────────
# 1. Prestashop vs Jabiru ES
with st.expander("📦 **Prestashop ↔ Jabiru ES**  — referencias PS que faltan en Jabiru (y viceversa)", expanded=True):
    ps_refs_upper = set(refs.str.strip().str.upper())
    # PS refs que no tienen base en Jabiru active (en PS hay algo que no está en Amazon ES)
    ps_bases_set = {extract_base(r).upper() for r in refs}
    missing_in_jabiru = pd.DataFrame([
        {"Referencia PS": r, "Base SKU": extract_base(r).upper()}
        for r in refs
        if extract_base(r).upper() not in {extract_base(s).upper() for s in jab_all["sku"]}
    ])
    # Jabiru active SKUs whose base is not in PS at all
    missing_in_ps_from_jabiru = pd.DataFrame([
        {"SKU Jabiru ES": sku, "Base SKU": b, "ASIN": asin}
        for b, (sku, asin) in j_bases.items()
        if b not in ps_refs_upper and f"S{b}" not in ps_refs_upper
    ])

    t1, t2 = st.tabs([
        f"PS sin match en Jabiru ({len(missing_in_jabiru)})",
        f"Jabiru sin match en PS ({len(missing_in_ps_from_jabiru)})"
    ])
    with t1:
        st.caption("Referencias de Prestashop que no tienen ninguna variante activa en Jabiru ES.")
        show_table(missing_in_jabiru, "ps_vs_jab", "ps_sin_jabiru.csv")
    with t2:
        st.caption("SKUs activos en Jabiru ES cuya base no aparece en Prestashop.")
        show_table(missing_in_ps_from_jabiru, "jab_vs_ps", "jabiru_sin_ps.csv")
        action_buttons(missing_in_ps_from_jabiru, "jab_vs_ps", "Jabiru→PS",
                       es_feed, es_feed, 1.0, sql_lk, carga_lk,
                       show_amz=False, show_ps=True)

st.divider()

# 2. Comparativa por tienda: Jabiru ES vs cada store
STORE_FEED = {
    "jabiru_fr": ("FR", "FR", 1.0),
    "jabiru_it": ("IT", "IT", 1.0),
    "jabiru_de": ("DE", "DE", 1.0),
    "jabiru_nl": ("DE", "",  1.0),
    "jabiru_pl": ("DE", "",  st.session_state.get("rate_pln",4.25)),
    "jabiru_se": ("DE", "",  st.session_state.get("rate_sek",11.5)),
    "turaco_es": ("ES", "",  1.0),
    "turaco_fr": ("FR", "FR",1.0),
    "turaco_it": ("IT", "IT",1.0),
    "turaco_de": ("DE", "DE",1.0),
}

prefixes_by_key = {k: STORE_CONFIG[k][2] for k in STORE_CONFIG}

for key, lbl, flag in OTHER_STORES:
    if not store_enabled[key]:
        continue
    f = store_files[key]
    label = lbl
    prefixes = prefixes_by_key[key]  # e.g. ["","S"] or ["","S","FR"]
    feed_key, _, rate = STORE_FEED[key]
    country_feed = feed_maps.get(feed_key, {})

    with st.expander(f"{flag} **Jabiru ES → {label}**", expanded=True):
        if f is None:
            st.info(f"Sube el listing de {label} para analizar.")
            continue
        with st.spinner(f"Leyendo {label}…"):
            _, store_all, store_attrs = parse_listing(f.read(), label)
        store_skus = sku_set(store_all)

        df_miss = missing_variants(j_bases, store_skus, prefixes)
        st.caption(f"{label}: **{store_attrs['active']:,}** activos / {store_attrs['total']:,} total")

        n_ok   = len(j_bases) - len(df_miss)
        n_miss = len(df_miss)
        c1, c2, c3 = st.columns(3)
        c1.metric("Bases Jabiru ES",  len(j_bases))
        c2.metric("✅ Con todas variantes", n_ok)
        c3.metric("❌ Con variantes faltantes", n_miss)

        show_table(df_miss, f"miss_{key}", f"faltantes_{key}.csv")
        if not df_miss.empty:
            action_buttons(df_miss, key, label,
                           country_feed, es_feed, rate,
                           sql_lk, carga_lk,
                           show_amz=(key in ("turaco_es",)),
                           show_ps=True)

    st.divider()

# 3. Espejos vs Prestashop
with st.expander("🔁 **Todos los espejos → Prestashop**  — variantes S+SKU / XX+SKU ausentes en PS", expanded=False):
    st.caption("Comprueba que Prestashop tiene creadas TODAS las referencias que existen en cualquier tienda Amazon.")
    df_ps_full = ps_missing(j_bases, {}, refs)
    m1, m2 = st.columns(2)
    m1.metric("Total variantes a comprobar", len(j_bases) * 6)  # base+S+FR+IT+DE ~ 5-6
    m2.metric("❌ Ausentes en PS", len(df_ps_full))
    show_table(df_ps_full, "ps_full", "faltantes_ps_completo.csv")
    if not df_ps_full.empty:
        action_buttons(df_ps_full, "ps_full", "PS completo",
                       es_feed, es_feed, 1.0, sql_lk, carga_lk,
                       show_amz=False, show_ps=True)
